"""
FixedRatePredictor

Gets files from the Bank of England website and charts the recent history of swap rates
Should predict movements in fixed rates but doesn't do that yet.
"""

# Standard library imports
import os
import zipfile
from calendar import monthrange
from datetime import date, datetime as dt
from math import ceil, floor
import pickle
import sqlite3
import sys
from time import localtime, strftime

# Third party imports
import certifi
import numpy as np
import pandas as pd
from sklearn.linear_model import LinearRegression
import slack
import tweepy
import urllib3
import yaml
from openpyxl import load_workbook
from pandas.plotting import register_matplotlib_converters
from pandas.tseries.offsets import BDay, BMonthBegin, BMonthEnd
from retry import retry

# have to import matplotlib separately first
# then change away from x-using backend
# then import pyplot
import matplotlib
register_matplotlib_converters()
matplotlib.use('Agg')
import matplotlib.pyplot as plt

# set the swap rate tenors we are interested in
TERMS = [2, 3, 4, 5, 10]

BANK_PATH = os.path.join('temp', 'yields', 'BLC Nominal daily data current month.xlsx')
GOV_PATH = os.path.join('temp', 'yields', 'GLC Nominal daily data current month.xlsx')
CHART_SAVE = 'chart.png'
DB_PATH = 'fixed_rate.db'
LOG_PATH = 'log.txt'
BG_COLOUR = '#FAFAFD'
FG_COLOUR = '#EEEEEE'
CHART_FONT = {'fontname': 'Cabin'}


# custom error in case BoE website not updated
class OutdatedFileError(Exception):
    """
    Bank of England file normally uploaded around midday, but sometimes isn't. This class prompts
    a retry of the file download if the file hasn't yet been updated.
    """

    pass


def build_prediction_model():
    boe = load_boe_history()
    shb = load_shb_history()
    # filter BOE data to match SHB date range
    shb_min = shb['valid_from'].min()
    shb_max = shb['valid_from'].max()
    mask = (boe['Date'] >= shb_min) & (boe['Date'] <= shb_max)
    boe = boe.loc[mask]

    # reduce BOE data to start-of-month values
    boe['fom'] = boe['Date'].apply(lambda x: BMonthBegin().rollback(x))
    boe = boe[boe['Date'] == boe['fom']]

    # add in monthly changes
    for term in TERMS:
        boe[f'{term}ydiff'] = boe[f'{term}y'].diff()
        shb[f'{term}diff'] = shb[term].diff()

    df = boe.merge(shb, on='period', how='inner')
    # drop the first row - which will contain NaN
    df = df.iloc[1:-1]
    for term in TERMS:
        X = df[f'{term}ydiff'].values.reshape(-1, 1)
        y = df[f'{term}diff'].values.reshape(-1, 1)
        model = LinearRegression().fit(X, y)
        r2 = model.score(X, y)
        store(term, model, 'sklearn_LinearRegression', r2)


def chart_rate_moves():
    boe = load_boe_history()
    shb = load_shb_history()
    # filter BOE data to match SHB date range
    shb_min = shb['valid_from'].min()
    shb_max = shb['valid_from'].max()
    mask = (boe['Date'] >= shb_min) & (boe['Date'] <= shb_max)
    boe = boe.loc[mask]

    # reduce BOE data to start-of-month values
    boe['fom'] = boe['Date'].apply(lambda x: BMonthBegin().rollback(x))
    boe = boe[boe['Date'] == boe['fom']]

    # add in monthly changes
    for rate in TERMS:
        boe[f'{rate}ydiff'] = boe[f'{rate}y'].diff()
        shb[f'{rate}diff'] = shb[rate].diff()

    df = boe.merge(shb, on='period', how='inner')

    rows = 1
    cols = 5
    fig, axs = plt.subplots(rows, cols,
                            figsize=(15, 4),
                            squeeze=False,
                            sharey=True)
    for idx, rate in enumerate(TERMS):
        row = idx % rows
        col = idx // rows
        ax = axs[row][col]
        residual = df[f'{rate}diff'] - df[f'{rate}ydiff']
        # ax.plot(df['Date'], residual, 'r+', label='residual')
        # ax.plot(boe['Date'], boe[f'{rate}ydiff'], 'g', label='boe diff')
        # ax.plot(shb['valid_from'], shb[f'{rate}diff'], 'c.', label='shb diff')
        ax.plot(df[f'{rate}ydiff'], df[f'{rate}diff'], 'b+')
        ax.set_title(f'{rate} year rate')
        ax.legend()
    plt.show()


@retry(OutdatedFileError, delay=60, backoff=5, max_delay=7500)
def daily_chart():
    write_log('Starting up')

    config = load_config()

    if config['download_file']:
        file_url = 'https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/latest-yield-curve-data.zip'
        file_name = get_file(file_url)
    else:
        file_name = os.path.join('yield.zip')

    proj_dir = os.path.dirname(os.path.realpath(__file__))
    zip_ref = zipfile.ZipFile(os.path.join(proj_dir, 'temp', file_name), 'r')
    zip_ref.extractall(os.path.join(proj_dir, 'temp', 'yields'))
    zip_ref.close()

    check_date = config['check_date']

    if check_date:
        workbook = load_workbook(filename=os.path.join(proj_dir, BANK_PATH))
        worksheet = workbook['4. spot curve']
        lastrow = str(workbook['4. spot curve'].max_row)
        filedate = worksheet['A' + lastrow].value.date()
        prevday = date.today() - BDay(1)

        if filedate != prevday.date():
            write_log('Date not OK, retrying...')
            raise OutdatedFileError

    bank_data = extract_data(BANK_PATH, '4. spot curve')

    bank_data.loc[:, 'Date'] = bank_data.loc[:, 'Date'].dt.day

    make_chart('Bank', bank_data)

    if config['send_tweet']:
        send_to_twitter(CHART_SAVE)
    if config['send_slack']:
        send_to_slack(CHART_SAVE)

    predict_rate_change(bank_data)

    write_log('Done\n----------------')


def db_connect():
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    return con, cur


def get_file(file_url):
    write_log('Downloading file')
    projdir = os.path.dirname(os.path.realpath(__file__))
    http = urllib3.PoolManager(
        cert_reqs='CERT_REQUIRED',
        ca_certs=certifi.where())
    r = http.request('GET', file_url)

    file_name = file_url.rsplit('/', 1)[-1]

    save_name = os.path.join(f'{date.today().strftime("%Y%m%d")} {file_name}')

    try:
        with open(os.path.join(projdir, 'temp', save_name), 'w+b') as f:
            f.write(r.data)
    except FileNotFoundError:
        os.mkdir(os.path.join(projdir, 'temp'))
        with open(os.path.join(projdir, 'temp', save_name), 'w+b') as f:
            f.write(r.data)

    return save_name


def extract_data(wb_path, sheet_name):
    """
    Returns a dataframe containing the data from a given path and sheet
    :param wb_path:
    :return:
    """
    projdir = os.path.dirname(os.path.realpath(__file__))
    worksheet = load_workbook(os.path.join(projdir, wb_path))[sheet_name]
    worksheet['A4'].value = "Date"

    # openpyxl max_row is unreliable with blank rows
    # so work out the max row manually
    lastrow = worksheet.max_row
    for r in range(lastrow, 1, -1):
        if worksheet.cell(row=r, column=1).value is not None:
            lastrow = r
            break

    # same for columns
    lastcol = worksheet.max_column
    for c in range(lastcol, 1, -1):
        if worksheet.cell(row=4, column=c).value is not None:
            lastcol = c
            break

    cols = [c.value for c in worksheet[4] if c.value is not None]
    data = list()
    for r in range(6, lastrow+1):
        rowdata = list()
        for c in range(1, lastcol+1):
            rowdata.append(worksheet.cell(row=r, column=c).value)
        data.append(rowdata)

    df_raw = pd.DataFrame(data, columns=cols)

    cols = TERMS.copy()
    cols.sort()
    df = df_raw[cols].dropna()
    df.columns = [f'{c}y' for c in cols]
    df /= 100
    df = df.assign(Date=df_raw.loc[:, 'Date'])
    return df


def load_config():
    proj_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(proj_dir, "config.yml")
    config = yaml.safe_load(open(config_path))
    return config


def load_boe_history():
    file_name = get_file('https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/blcnomddata.zip')
    proj_dir = os.path.dirname(os.path.realpath(__file__))
    zip_folder = os.path.join(proj_dir, 'temp', 'yield-archive')
    zip_ref = zipfile.ZipFile(os.path.join(proj_dir, 'temp', file_name), 'r')
    try:
        zip_ref.extractall(zip_folder)
    except FileNotFoundError:
        os.mkdir(zip_folder)
        zip_ref.extractall(zip_folder)

    zip_ref.close()

    wb_file = os.path.join(zip_folder, 'BLC Nominal daily data_2016 to present.xlsx')
    boe_hist = extract_data(wb_file, '4. spot curve')
    boe_hist['period'] = boe_hist['Date'].dt.strftime('%y%m')
    return boe_hist


def load_shb_history():
    proj_dir = os.path.dirname(os.path.abspath(__file__))
    s = pd.read_csv(os.path.join(proj_dir, 'model', 'fixed_rate_history.csv'),
                    header=0,
                    parse_dates=['valid_from'],
                    dayfirst=True)
    # todo: work with mid-month updates
    s = s.loc[s['update_type'] == 'start_of_month']
    s = s[['valid_from', 'rmc_rate', 'fix_length']]
    r = s.pivot(index='valid_from', columns='fix_length', values='rmc_rate')
    # limit to 2016 onwards to match SHB data
    r = r.loc['2016-01-01':]
    r = r.reset_index()
    r['period'] = r['valid_from'].dt.strftime('%y%m')
    return r


def make_chart(df_name, df):
    """
    Creates a chart from one input dataframes, and saves it to a PNG file.
    TODO: return filename, and merge with make_charts(?)
    :param df_name: name of the data being used (e.g. bank, sovereign)
    :param df: pandas dataframe containing yield curve data
    :return: none
    '''
    """
    write_log('Making chart')
    projdir = os.path.dirname(os.path.realpath(__file__))

    colours = ['#00E87E', '#0C87F2', '#5000DB', '#F20F7B', '#E85200', '#FF6B26']
    if len(TERMS) > len(colours):
        print("Falling back to colormap")
        colours = plt.cm.Accent(np.linspace(0, 1, len(TERMS) + 1))

    fig, ax = plt.subplots(figsize=(5, 4))

    # set up chart format
    ax.set_facecolor(BG_COLOUR)
    ax.set_prop_cycle('color', colours)
    ax.grid(color=FG_COLOUR, linestyle='-', linewidth=0.5)

    # work out the start and end dates of the month, and format x axis accordingly
    dmin = df.iloc[0].loc['Date']
    dmax = monthrange(date.today().year, date.today().month)[1]
    ax.set_xlim(1, dmax)

    # plot the path of rates for each term
    cols = [f'{t}y' for t in TERMS]
    ax.plot(df.loc[:, 'Date'], df.loc[:, cols])

    # format axis labels
    plt.title(df_name, **CHART_FONT)

    ax.set_yticklabels((f'{y*100 : 1.2f}%' for y in ax.get_yticks()),
                       **CHART_FONT)
    # todo: next line doesn't seem to have any effect
    ax.set_ybound(floor(ax.get_ybound()[0] * 1000) / 1000,
                  ceil(ax.get_ybound()[1] * 1000) / 1000)
    ax.set_xticklabels((f'{x : 1.0f}' for x in ax.get_xticks()),
                       **CHART_FONT)
    ax.set_xlabel("Day", **CHART_FONT)

    ymin, ymax = ax.get_ylim()
    yrange = ymax - ymin

    drpt = df.iloc[-1].loc['Date']
    today = date.today()
    dmax = monthrange(today.year, today.month)[1]

    for j, col in enumerate(df.loc[:, cols]):
        start_rate = df.iloc[0].loc[col]
        # plot a dashed line showing the start-of-month value for each term
        ax.plot((dmin, dmax), (start_rate, start_rate),
                linestyle=":", linewidth=1, color=colours[j])
        # label end of dashed line with rate from day one and the relevant term e.g. 2yr, 10yr
        ax.annotate(f'  {start_rate*100:1.2f}%  {col}r',
                    xy=(dmax, df.iloc[0].loc[col] - 0.015 * yrange),
                    xycoords='data',
                    color=colours[j],
                    fontsize=10,
                    **CHART_FONT)
        # label end of plotted line with current rate
        # first, work out if displacement needed to avoid clash
        labeloffset = 0
        latest_rate = df.iloc[-1].loc[col]
        rate_diff = abs(latest_rate-start_rate)

        if rate_diff < 0.0002:
            labeloffset = -0.0002

        ax.annotate(f'  {latest_rate*100:1.2f}%',
                    xy=(drpt, df.iloc[-1].loc[col] + labeloffset),
                    xycoords='data',
                    color=colours[j],
                    fontsize=10,
                    **CHART_FONT)

    # plt.tight_layout(rect=[0.9, 0.9, 0.9, 0.9])
    plt.suptitle(strftime('Swap rates, %b %Y', localtime()), y=0.98, **CHART_FONT)
    plt.savefig(os.path.join(projdir, CHART_SAVE), facecolor=BG_COLOUR, edgecolor='none')


def make_charts(dfs):
    """
    Creates a chart from two input dataframes, and saves it to a PNG file.
    TODO: return filename
    :param dfs:
    :return: none
    '''
    """
    write_log('Making chart')
    projdir = os.path.dirname(os.path.realpath(__file__))

    colours = plt.cm.Set2(np.linspace(0, 1, len(TERMS)+1))

    fig = plt.figure(figsize=(8, 4))

    axs = list()

    for i, data in enumerate(dfs, 1):
        dfname = data[0]
        df = data[1]
        x, y = (1, 2)
        if i % 2 == 0:
            ax1 = axs[-1]
            axs.append(fig.add_subplot(x, y, i, sharey=ax1))
            plt.setp(axs[-1].get_yticklabels(), visible=False)
        else:
            axs.append(fig.add_subplot(x, y, i))

        # set up chart format
        axs[-1].set_facecolor(BG_COLOUR)
        axs[-1].set_prop_cycle('color', colours)
        axs[-1].grid(color=FG_COLOUR, linestyle='-', linewidth=0.5)

        # work out the start and end dates of the month, and format x axis accordingly
        dmin = df.iloc[0].loc['Date']
        dmax = monthrange(date.today().year, date.today().month)[1]
        axs[-1].set_xlim(1, dmax)

        axs[-1].plot(df.loc[:, 'Date'], df.loc[:, TERMS])

        # plot a dashed line showing the start-of-month value for each term
        for j, col in enumerate(df.loc[:, TERMS]):
            axs[-1].plot((dmin, dmax), (df.iloc[0].loc[col], df.iloc[0].loc[col]),
                         linestyle=":", linewidth=1)

        # format axis labels
        plt.title(dfname, **CHART_FONT)
        axs[-1].set_ybound(floor(axs[-1].get_ybound()[0] * 1000) / 1000,
                           ceil(axs[-1].get_ybound()[1] * 1000) / 1000)
        axs[-1].set_yticklabels((f'{x*100 : 1.2f}%' for x in axs[-1].get_yticks()),
                                **CHART_FONT)
        axs[-1].set_xticklabels((f'{x : 1.0f}' for x in axs[-1].get_xticks()),
                                **CHART_FONT)
        axs[-1].set_xlabel("Day", **CHART_FONT)

    ymin, ymax = axs[-1].get_ylim()
    yrange = ymax - ymin

    for i, data in enumerate(dfs):
        df = data[1]
        ax = axs[i]
        dmin = df.index.values[0]
        drpt = df.index.values[-1]
        today = date.today()
        dmax = monthrange(today.year, today.month)[1]
        for j, col in enumerate(df):
            # label near end of dashed line with relevant term (2yr, 10yr etc)
            ax.annotate(col + 'r',
                        xy=(dmax - 0.5, df.iloc[0].loc[col] + 0.0150 * yrange),
                        xycoords='data',
                        ha='right',
                        color=colours[j],
                        fontsize=12,
                        **CHART_FONT)
            # label end of dashed line with rate from day one
            ax.annotate('  {:1.2f}%'.format(100 * df.loc[dmin, col]),
                        xy=(dmax, df.iloc[0].loc[col] - 0.015 * yrange),
                        xycoords='data',
                        color=colours[j],
                        fontsize=10,
                        **CHART_FONT)
            # label end of plotted line with current rate
            # first, work out if displacement needed to avoid clash
            labeloffset = 0
            ratediff = df.loc[drpt, col] - df.loc[dmin, col]

            if abs(ratediff) < 0.0002:
                labeloffset = -0.0002
            ax.annotate('  {:1.2f}%'.format(100 * df.loc[drpt, col]),
                        xy=(drpt, df.loc[drpt, col] + labeloffset),
                        xycoords='data',
                        color=colours[j],
                        fontsize=10,
                        **CHART_FONT)

    plt.tight_layout(rect=[-0.010, 0, 0.85 + 0.05 * len(dfs), 0.94])
    plt.subplots_adjust(wspace=0.20)
    plt.suptitle(strftime('Swap rates, %b %Y', localtime()), y=0.98, **CHART_FONT)
    plt.savefig(os.path.join(projdir, CHART_SAVE), facecolor=BG_COLOUR, edgecolor='none')


def predict_rate_change(data):
    messages = []
    for term in TERMS:
        opening_rate = data[f'{term}y'].iloc[0]
        closing_rate = data[f'{term}y'].iloc[-1]
        # basic prediction model
        rate_change = closing_rate - opening_rate
        _, curs = db_connect()
        pickled_model = curs.execute('''SELECT pickle 
                                        FROM model 
                                        WHERE rate_term = ?
                                        ORDER BY r_2 DESC
                                        LIMIT 1''', (term,)).fetchone()
        model = pickle.loads(pickled_model[0])
        prediction = model.predict([[rate_change]])
        predicted_change = round_nearest(prediction[0][0], 0.0005)

        # month_end = BMonthEnd().rollforward(date.today())

        msg = f'{term}\t\t{rate_change:.4%}\t{predicted_change:.2%}'

        messages.append(msg)

    if messages:
        messages.insert(0, 'term\tchange\t\tprediction')
        msg_text = '\n'.join(messages)
        print(msg_text)
        config = load_config()
        slack_token = config['slack_login']['bot_token']
        client = slack.WebClient(token=slack_token)
        response = client.chat_postMessage(channel=config['slack_channel'],
                                           text=msg_text)
        assert response["ok"]


def round_nearest(x, a):
    return round(x / a) * a


def send_to_slack(imgpath):
    write_log('Sending plot to Slack')
    proj_dir = os.path.dirname(os.path.realpath(__file__))
    file_path = os.path.join(proj_dir, imgpath)

    config = load_config()
    slack_token = config['slack_login']['bot_token']
    client = slack.WebClient(token=slack_token)

    response = client.files_upload(
        channels=config['slack_channel'],
        file=file_path)
    assert response["ok"]


def send_to_twitter(imgpath):
    write_log('Tweeting plot')
    projdir = os.path.dirname(os.path.realpath(__file__))

    config = load_config()
    twit_auth = config['twitter_login']

    auth = tweepy.OAuthHandler(twit_auth['consumer_key'], twit_auth['consumer_secret'])
    auth.set_access_token(twit_auth['access_key'], twit_auth['access_secret'])
    api = tweepy.API(auth)
    api.update_with_media(os.path.join(projdir, imgpath))


def store(term, model_obj, model_type, r_2):
    conn, curs = db_connect()
    curs.execute('''CREATE TABLE IF NOT EXISTS
                    model (
                    id PRIMARY KEY,
                    pickle BLOB,
                    model_type TEXT,
                    date_stored TEXT,
                    rate_term REAL,
                    r_2 REAL)''')
    pickled_model = pickle.dumps(model_obj, protocol=2)
    values = (sqlite3.Binary(pickled_model),
              model_type,
              dt.now(),
              term,
              r_2)

    curs.execute('''INSERT INTO model
                    (pickle, model_type, date_stored, rate_term, r_2)
                    VALUES (?, ?, ?, ?, ?)''', values)
    conn.commit()


def write_log(log_text):
    projdir = os.path.dirname(os.path.realpath(__file__))
    with open(os.path.join(projdir, LOG_PATH), 'a') as f:
        log_time = strftime('%Y-%m-%d %H:%M:%S', localtime())
        f.write(f'{log_time} {log_text}\n')


if __name__ == '__main__':
    environment = load_config()['mode']
    if environment == "production":
        if len(sys.argv) != 2:
            raise SyntaxError('FRP takes exactly one argument.')

        mode = sys.argv[1]
        if mode == 'd':
            daily_chart()
        elif mode == 'm':
            build_prediction_model()
        else:
            raise SyntaxError('Valid modes are d (produce daily chart) or m (build prediction model)')

    elif environment == "development":
        daily_chart()

    else:
        print(f"Mode ({environment}) specified in config.yml is invalid")
